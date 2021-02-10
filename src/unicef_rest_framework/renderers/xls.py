import logging
import re
from collections.abc import Iterable, MutableMapping

from drf_renderer_xlsx.renderers import XLSXRenderer as _XLSXRenderer
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError

from unicef_rest_framework.renderers.mixin import ContentDispositionMixin

logger = logging.getLogger(__name__)


class XLSXRenderer(ContentDispositionMixin, _XLSXRenderer):

    def _flatten(self, data, parent_key='', key_sep='.', list_sep=', '):
        items = []
        for k, v in data.items():
            new_key = f"{parent_key}{key_sep}{k}" if parent_key else k
            if isinstance(v, MutableMapping):
                items.extend(self._flatten(v, new_key, key_sep=key_sep).items())
            elif isinstance(v, Iterable) and not isinstance(v, str):
                # Flatten the array into a comma separated string to fit
                # in a single spreadsheet column
                vv = []
                for item in v:
                    if isinstance(item, dict):
                        vv.append("\n".join([f"{ki}: {vi}" for ki, vi in item.items() if vi]))
                    elif isinstance(item, list):
                        vv.append(': '.join(item))
                    else:
                        vv.append(item)
                items.append((new_key, list_sep.join(vv)))
            else:
                items.append((new_key, v))
        return dict(items)

    def _make_body(self, row, row_count):
        column_count = 0
        row_count += 1
        flatten_row = self._flatten(row)
        for column_name, value in flatten_row.items():
            if column_name == "row_color":
                continue
            column_count += 1
            try:
                cell = self.ws.cell(row=row_count, column=column_count, value=value)
            except IllegalCharacterError:
                value = re.sub(r'\W+', ' ', value)
                cell = self.ws.cell(row=row_count, column=column_count, value=value)
            cell.style = self.body_style
        self.ws.row_dimensions[row_count].height = self.body.get("height", 40)
        if "row_color" in row:
            last_letter = get_column_letter(column_count)
            cell_range = self.ws[
                "A{}".format(row_count): "{}{}".format(last_letter, row_count)
            ]
            fill = PatternFill(fill_type="solid", start_color=row["row_color"])
            for r in cell_range:
                for c in r:
                    c.fill = fill

    def render(self, data, accepted_media_type=None, renderer_context=None):
        response = renderer_context['response']
        self.process_response(renderer_context)
        if response.status_code != 200:
            return ''
        try:
            return super().render(data, accepted_media_type, renderer_context)

        except Exception as e:
            logger.exception(e)
            raise Exception(f'Error processing request {e}') from e
